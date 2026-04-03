"""
DCF Excel Export — generates .xlsx workbook with live formulas.

Sheets: Summary | P&L | DCF | Sensitivity | Comps (if data)
Convention: Blue font = input, Black = formula, Green = output.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Style constants (IB convention) ──────────────────────────
BLUE_FONT = Font(name="Calibri", size=10, color="0066CC")          # Input
BLACK_FONT = Font(name="Calibri", size=10, color="000000")         # Formula
GREEN_FONT = Font(name="Calibri", size=10, color="006100", bold=True)  # Output
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=10, bold=True)
LIGHT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PCT_FMT = "0.0%"
NUM_FMT = "#,##0.0"
DOLLAR_FMT = "$#,##0.00"


# ── Public API ───────────────────────────────────────────────

def export_dcf_excel(
    dcf_result: dict,
    assumptions: dict,
    output_path: str,
    comps_data: dict | None = None,
) -> str:
    """Generate Excel workbook with live formulas. Returns output_path."""
    wb = Workbook()

    yby = dcf_result.get("year_by_year", [])
    n_years = len(yby)

    _build_summary_sheet(wb, dcf_result, assumptions)
    _build_pl_sheet(wb, yby, assumptions, n_years)
    _build_dcf_sheet(wb, dcf_result, assumptions, n_years)
    _build_sensitivity_sheet(wb, dcf_result)
    if comps_data:
        _build_comps_sheet(wb, comps_data)

    # Remove default empty sheet if still present
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)
    return output_path


# ── Helpers ──────────────────────────────────────────────────

def _write_header_row(ws, row: int, labels: list, start_col: int = 1):
    for i, label in enumerate(labels):
        cell = ws.cell(row=row, column=start_col + i, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _write_row(ws, row: int, col: int, values: list, font=BLACK_FONT, fmt=None):
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=col + i, value=v)
        cell.font = font
        if fmt:
            cell.number_format = fmt


# ── Sheet builders ───────────────────────────────────────────

def _build_summary_sheet(wb: Workbook, dcf_result: dict, assumptions: dict):
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    # Title
    company = assumptions.get("company", "")
    ticker = assumptions.get("ticker", "")
    ws.cell(row=1, column=1, value=f"{company} ({ticker}) — DCF Summary")
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True)

    # Key outputs (green)
    r = 3
    for label, key, fmt in [
        ("Fair Value per Share", "fair_value_per_share", DOLLAR_FMT),
        ("Current Price", "current_price", DOLLAR_FMT),
        ("Upside / Downside", "gap_pct", "0.0%"),
        ("Enterprise Value ($M)", "enterprise_value", NUM_FMT),
        ("Equity Value ($M)", "equity_value", NUM_FMT),
    ]:
        ws.cell(row=r, column=1, value=label).font = SECTION_FONT
        val = dcf_result.get(key)
        if key == "gap_pct" and val is not None:
            val = val / 100
        cell = ws.cell(row=r, column=2, value=val)
        cell.font = GREEN_FONT
        cell.number_format = fmt
        r += 1

    # Key assumptions (blue = input)
    r += 1
    ws.cell(row=r, column=1, value="Key Assumptions").font = Font(
        name="Calibri", size=12, bold=True
    )
    r += 1
    assumptions_items = [
        ("WACC", assumptions.get("wacc"), PCT_FMT),
        ("Terminal Growth", assumptions.get("terminal_growth"), PCT_FMT),
        ("Projection Years", assumptions.get("projection_years"), "0"),
        ("Shares Outstanding", assumptions.get("shares_outstanding"), "#,##0"),
        ("Net Cash ($M)", assumptions.get("net_cash"), NUM_FMT),
    ]
    gm = assumptions.get("gross_margin")
    if isinstance(gm, dict):
        assumptions_items.append(("Gross Margin", gm.get("value"), PCT_FMT))

    for label, val, fmt in assumptions_items:
        ws.cell(row=r, column=1, value=label).font = BLACK_FONT
        cell = ws.cell(row=r, column=2, value=val)
        cell.font = BLUE_FONT
        cell.number_format = fmt
        r += 1

    # Sell-side anchor if present
    anchor = dcf_result.get("sell_side_anchor")
    if anchor:
        r += 1
        ws.cell(
            row=r, column=1,
            value=f"Sell-Side Anchor ({anchor.get('source', '?')})",
        ).font = Font(name="Calibri", size=12, bold=True)
        r += 1
        for k, v in anchor.items():
            if k == "source":
                continue
            ws.cell(row=r, column=1, value=k).font = BLACK_FONT
            ws.cell(row=r, column=2, value=v).font = Font(
                name="Calibri", size=10, color="996600"
            )
            r += 1


def _build_pl_sheet(wb: Workbook, yby: list, assumptions: dict, n_years: int):
    ws = wb.create_sheet("P&L")
    ws.column_dimensions["A"].width = 24

    headers = [""] + [f"Y{row['year']}" for row in yby]
    _write_header_row(ws, 1, headers)
    for i in range(1, len(headers)):
        ws.column_dimensions[get_column_letter(i + 1)].width = 16

    rows_data: list[tuple[str, list, Font, str]] = [
        ("Revenue ($M)", [r["revenue"] for r in yby], BLUE_FONT, NUM_FMT),
        ("Revenue Growth", [r["revenue_growth"] for r in yby], BLACK_FONT, PCT_FMT),
        ("COGS ($M)", [r.get("cogs", r["revenue"] - r["gross_profit"]) for r in yby], BLACK_FONT, NUM_FMT),
        ("Gross Profit ($M)", [r["gross_profit"] for r in yby], BLACK_FONT, NUM_FMT),
    ]

    # Optional SGA/R&D detail rows
    if yby and "sga" in yby[0]:
        rows_data.append(("SG&A ($M)", [r.get("sga", 0) for r in yby], BLACK_FONT, NUM_FMT))
    if yby and "rd" in yby[0]:
        rows_data.append(("R&D ($M)", [r.get("rd", 0) for r in yby], BLACK_FONT, NUM_FMT))

    rows_data += [
        ("Total OpEx ($M)", [r.get("total_opex", r["gross_profit"] - r["ebit"]) for r in yby], BLACK_FONT, NUM_FMT),
        ("EBIT ($M)", [r["ebit"] for r in yby], SECTION_FONT, NUM_FMT),
        ("NOPAT ($M)", [r["nopat"] for r in yby], BLACK_FONT, NUM_FMT),
    ]

    r = 2
    rev_row = r  # track revenue row for formulas
    gp_row = None
    for label, values, font, fmt in rows_data:
        ws.cell(row=r, column=1, value=label).font = SECTION_FONT
        _write_row(ws, r, 2, values, font, fmt)
        if "Gross Profit" in label:
            gp_row = r
        r += 1

    # Add formula: Gross Margin % = Gross Profit / Revenue
    if gp_row and n_years > 0:
        r += 1
        ws.cell(row=r, column=1, value="Gross Margin %").font = BLACK_FONT
        for i in range(n_years):
            col_letter = get_column_letter(i + 2)
            cell = ws.cell(
                row=r, column=i + 2,
                value=f"={col_letter}{gp_row}/{col_letter}{rev_row}",
            )
            cell.font = BLACK_FONT
            cell.number_format = PCT_FMT


def _build_dcf_sheet(wb: Workbook, dcf_result: dict, assumptions: dict, n_years: int):
    ws = wb.create_sheet("DCF")
    ws.column_dimensions["A"].width = 24

    yby = dcf_result.get("year_by_year", [])
    headers = [""] + [f"Y{row['year']}" for row in yby] + ["Terminal"]
    _write_header_row(ws, 1, headers)
    for i in range(1, len(headers)):
        ws.column_dimensions[get_column_letter(i + 1)].width = 16

    wacc = assumptions.get("wacc", 0.10)

    rows_data = [
        ("NOPAT ($M)", [r["nopat"] for r in yby]),
        ("+ D&A ($M)", [r["da"] for r in yby]),
        ("- CapEx ($M)", [r.get("capex", 0) for r in yby]),
        ("\u0394WC ($M)", [r.get("delta_wc", 0) for r in yby]),
        ("= FCF ($M)", [r["fcf"] for r in yby]),
        ("Discount Factor", [(1 + wacc) ** (t + 1) for t in range(n_years)]),
        ("Discounted FCF ($M)", [r["discounted_fcf"] for r in yby]),
    ]

    r = 2
    nopat_row = r
    da_row = r + 1
    capex_row = r + 2
    dwc_row = r + 3
    fcf_row = r + 4
    df_row = r + 5

    for label, values in rows_data:
        ws.cell(row=r, column=1, value=label).font = SECTION_FONT
        is_fcf = label.startswith("= FCF")
        font = GREEN_FONT if is_fcf else BLACK_FONT
        _write_row(ws, r, 2, values, font, NUM_FMT)
        r += 1

    # Overwrite FCF cells with live formulas: =NOPAT + D&A - CapEx - dWC
    for i in range(n_years):
        cl = get_column_letter(i + 2)
        cell = ws.cell(
            row=fcf_row, column=i + 2,
            value=f"={cl}{nopat_row}+{cl}{da_row}-{cl}{capex_row}-{cl}{dwc_row}",
        )
        cell.font = GREEN_FONT
        cell.number_format = NUM_FMT

    # Overwrite Discounted FCF with formula: =FCF/DiscountFactor
    for i in range(n_years):
        cl = get_column_letter(i + 2)
        cell = ws.cell(
            row=fcf_row + 2, column=i + 2,
            value=f"={cl}{fcf_row}/{cl}{df_row}",
        )
        cell.font = BLACK_FONT
        cell.number_format = NUM_FMT

    # Terminal value
    tv = dcf_result.get("terminal_value", 0)
    dtv = dcf_result.get("discounted_terminal_value", 0)
    r += 1
    ws.cell(row=r, column=1, value="Terminal Value ($M)").font = SECTION_FONT
    ws.cell(row=r, column=n_years + 2, value=tv).font = BLACK_FONT
    ws.cell(row=r, column=n_years + 2).number_format = NUM_FMT
    r += 1
    ws.cell(row=r, column=1, value="Discounted TV ($M)").font = SECTION_FONT
    ws.cell(row=r, column=n_years + 2, value=dtv).font = BLACK_FONT
    ws.cell(row=r, column=n_years + 2).number_format = NUM_FMT

    # Bridge to equity value
    r += 2
    sum_dcf = sum(row["discounted_fcf"] for row in yby)
    bridge = [
        ("Sum of Discounted FCF", sum_dcf, NUM_FMT),
        ("+ Discounted Terminal Value", dtv, NUM_FMT),
        ("= Enterprise Value ($M)", dcf_result.get("enterprise_value", 0), NUM_FMT),
        ("+ Net Cash ($M)", assumptions.get("net_cash", 0), NUM_FMT),
        ("= Equity Value ($M)", dcf_result.get("equity_value", 0), NUM_FMT),
        ("Shares Outstanding", assumptions.get("shares_outstanding", 0), "#,##0"),
        ("= Fair Value per Share", dcf_result.get("fair_value_per_share", 0), DOLLAR_FMT),
    ]
    bridge_start = r
    for label, val, fmt in bridge:
        ws.cell(row=r, column=1, value=label).font = SECTION_FONT
        is_output = label.startswith("= ")
        cell = ws.cell(row=r, column=2, value=val)
        cell.font = GREEN_FONT if is_output else BLACK_FONT
        cell.number_format = fmt
        r += 1

    # Live formula: EV = Sum DCF + Discounted TV
    ev_row = bridge_start + 2
    ws.cell(row=ev_row, column=2, value=f"=B{bridge_start}+B{bridge_start+1}")
    ws.cell(row=ev_row, column=2).font = GREEN_FONT
    ws.cell(row=ev_row, column=2).number_format = NUM_FMT

    # Live formula: Equity Value = EV + Net Cash
    eq_row = bridge_start + 4
    ws.cell(row=eq_row, column=2, value=f"=B{ev_row}+B{ev_row+1}")
    ws.cell(row=eq_row, column=2).font = GREEN_FONT
    ws.cell(row=eq_row, column=2).number_format = NUM_FMT

    # Live formula: Fair Value = Equity Value * 1000000 / Shares
    fv_row = bridge_start + 6
    shares_row = bridge_start + 5
    ws.cell(row=fv_row, column=2, value=f"=B{eq_row}*1000000/B{shares_row}")
    ws.cell(row=fv_row, column=2).font = GREEN_FONT
    ws.cell(row=fv_row, column=2).number_format = DOLLAR_FMT


def _build_sensitivity_sheet(wb: Workbook, dcf_result: dict):
    ws = wb.create_sheet("Sensitivity")
    sens = dcf_result.get("sensitivity", {})
    if not sens:
        ws.cell(row=1, column=1, value="No sensitivity data")
        return

    wacc_vals = sens.get("wacc_values", [])
    growth_vals = sens.get("growth_values", [])
    matrix = sens.get("matrix", [])

    ws.column_dimensions["A"].width = 14

    ws.cell(row=1, column=1, value="Fair Value Sensitivity: WACC vs Terminal Growth")
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=12, bold=True)

    # Column headers (terminal growth)
    ws.cell(row=3, column=1, value="WACC \\ TG").font = HEADER_FONT
    ws.cell(row=3, column=1).fill = HEADER_FILL
    for j, g in enumerate(growth_vals):
        cell = ws.cell(row=3, column=j + 2, value=g)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.number_format = PCT_FMT
        ws.column_dimensions[get_column_letter(j + 2)].width = 14

    # Data rows
    mid_r = len(wacc_vals) // 2
    mid_c = len(growth_vals) // 2
    for i, w in enumerate(wacc_vals):
        cell = ws.cell(row=4 + i, column=1, value=w)
        cell.font = SECTION_FONT
        cell.number_format = PCT_FMT
        for j in range(len(growth_vals)):
            val = matrix[i][j] if i < len(matrix) and j < len(matrix[i]) else None
            cell = ws.cell(row=4 + i, column=j + 2, value=val)
            cell.number_format = DOLLAR_FMT
            if i == mid_r and j == mid_c:
                cell.font = GREEN_FONT
                cell.fill = GREEN_FILL
            else:
                cell.font = BLACK_FONT


def _build_comps_sheet(wb: Workbook, comps_data: dict):
    ws = wb.create_sheet("Comps")
    peers = comps_data.get("peers", [])
    if not peers:
        ws.cell(row=1, column=1, value="No comps data")
        return

    headers = ["Ticker", "Market Cap ($M)", "Fwd P/E", "EV/EBITDA",
               "Rev Growth", "Gross Margin", "Target?"]
    _write_header_row(ws, 1, headers)
    ws.column_dimensions["A"].width = 12
    for i in range(2, 8):
        ws.column_dimensions[get_column_letter(i)].width = 16

    for i, p in enumerate(peers):
        r = i + 2
        ws.cell(row=r, column=1, value=p.get("ticker", "")).font = SECTION_FONT
        ws.cell(row=r, column=2, value=p.get("market_cap")).number_format = NUM_FMT
        ws.cell(row=r, column=3, value=p.get("fwd_pe")).number_format = "0.0x"
        ws.cell(row=r, column=4, value=p.get("ev_ebitda")).number_format = "0.0x"
        ws.cell(row=r, column=5, value=p.get("revenue_growth")).number_format = PCT_FMT
        ws.cell(row=r, column=6, value=p.get("gross_margin")).number_format = PCT_FMT
        is_target = p.get("is_target", False)
        ws.cell(row=r, column=7, value="\u2713" if is_target else "")
        if is_target:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = LIGHT_FILL
