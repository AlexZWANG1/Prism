"""Tests for enhanced Excel parsing with formula and structure extraction."""
import pytest
from openpyxl import Workbook
from tools.document_parser import parse_file


def _make_test_xlsx(path, with_formulas=True):
    wb = Workbook()
    ws = wb.active
    ws.title = "DCF Model"
    ws["A1"] = "Revenue"
    ws["B1"] = 1000
    ws["C1"] = "=B1*1.1" if with_formulas else 1100
    ws["A2"] = "WACC"
    ws["B2"] = 0.10
    wb.save(str(path))


def test_excel_parse_returns_markdown(tmp_path):
    path = tmp_path / "test.xlsx"
    _make_test_xlsx(path, with_formulas=False)
    result = parse_file(path.read_bytes(), path.name)
    assert "Revenue" in result.content
    assert "1000" in result.content


def test_excel_parse_extracts_formulas(tmp_path):
    path = tmp_path / "test.xlsx"
    _make_test_xlsx(path, with_formulas=True)
    result = parse_file(path.read_bytes(), path.name)
    assert "formula" in result.content.lower() or "=B1*1.1" in result.content


def test_excel_parse_preserves_sheet_names(tmp_path):
    path = tmp_path / "test.xlsx"
    _make_test_xlsx(path)
    result = parse_file(path.read_bytes(), path.name)
    assert "DCF Model" in result.content
