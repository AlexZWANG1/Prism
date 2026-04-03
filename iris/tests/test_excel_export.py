"""Tests for DCF Excel export with live formulas."""
import pytest
from skills.dcf.tools import build_dcf
from skills.dcf.excel_export import export_dcf_excel

BASE = {
    "company": "TestCo", "ticker": "TEST", "projection_years": 3,
    "segments": [{"name": "Main", "current_annual_revenue": 10000,
                  "growth_rates": [0.10, 0.08, 0.06]}],
    "gross_margin": {"value": 0.60},
    "opex_pct_of_revenue": {"value": 0.20},
    "wacc": 0.10, "terminal_growth": 0.03,
    "net_cash": 500, "shares_outstanding": 100_000_000,
    "current_price": 50.0,
}


def test_export_creates_file(tmp_path):
    result = build_dcf(BASE)
    assert result.status == "ok"
    path = tmp_path / "test_dcf.xlsx"
    export_dcf_excel(result.data, BASE, str(path))
    assert path.exists()
    assert path.stat().st_size > 5000  # non-trivial file


def test_export_has_all_sheets(tmp_path):
    result = build_dcf(BASE)
    path = tmp_path / "test_dcf.xlsx"
    export_dcf_excel(result.data, BASE, str(path))
    from openpyxl import load_workbook
    wb = load_workbook(str(path))
    names = wb.sheetnames
    assert "Summary" in names
    assert "P&L" in names
    assert "DCF" in names
    assert "Sensitivity" in names


def test_export_has_live_formulas(tmp_path):
    result = build_dcf(BASE)
    path = tmp_path / "test_dcf.xlsx"
    export_dcf_excel(result.data, BASE, str(path))
    from openpyxl import load_workbook
    wb = load_workbook(str(path))
    dcf_ws = wb["DCF"]
    has_formula = False
    for row in dcf_ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                has_formula = True
                break
    assert has_formula, "DCF sheet should have live formulas"


def test_export_blue_black_green_colors(tmp_path):
    result = build_dcf(BASE)
    path = tmp_path / "test_dcf.xlsx"
    export_dcf_excel(result.data, BASE, str(path))
    from openpyxl import load_workbook
    wb = load_workbook(str(path))
    assumptions_ws = wb["Summary"]
    found_blue = False
    for row in assumptions_ws.iter_rows():
        for cell in row:
            if cell.font and cell.font.color and cell.font.color.rgb:
                rgb = str(cell.font.color.rgb)
                if "0066CC" in rgb.upper():
                    found_blue = True
    assert found_blue, "Input cells should use blue font"
