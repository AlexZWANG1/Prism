"""
Document parser — multi-engine document extraction for PDF, Excel, and text files.

Architecture inspired by OpenNotebook (lfnovo/open-notebook):
  - L1 (fast/free): pymupdf4llm for text + basic table extraction
  - L2 (precise/free): Docling (IBM) for high-accuracy table parsing
  - Fallback: PyPDF2 if neither is available

All engines output Markdown, which preserves table structure and is
ideal for LLM consumption and downstream chunking.
"""

from __future__ import annotations

import io
import logging
import tempfile
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path

logger = logging.getLogger(__name__)


class ParseEngine(str, Enum):
    AUTO = "auto"           # Try best available engine
    PYMUPDF = "pymupdf"    # Fast, good for text-heavy PDFs
    DOCLING = "docling"     # Slow but 97.9% table accuracy
    PYPDF2 = "pypdf2"      # Fallback, text-only


@dataclass
class ParseResult:
    """Result of document parsing."""
    content: str                  # Extracted text (Markdown format)
    engine_used: str              # Which engine actually ran
    page_count: int = 0
    metadata: dict = field(default_factory=dict)  # title, author, etc.
    warnings: list[str] = field(default_factory=list)


# ── Engine availability detection ────────────────────────────

def _has_pymupdf() -> bool:
    try:
        import pymupdf4llm  # noqa: F401
        return True
    except ImportError:
        return False


def _has_docling() -> bool:
    try:
        from docling.document_converter import DocumentConverter  # noqa: F401
        return True
    except ImportError:
        return False


def _has_pypdf2() -> bool:
    try:
        from PyPDF2 import PdfReader  # noqa: F401
        return True
    except ImportError:
        return False


def available_engines() -> list[str]:
    """Return list of installed parsing engines."""
    engines = []
    if _has_pymupdf():
        engines.append("pymupdf")
    if _has_docling():
        engines.append("docling")
    if _has_pypdf2():
        engines.append("pypdf2")
    return engines


# ── PDF Parsing ──────────────────────────────────────────────

def _parse_pdf_pymupdf(pdf_bytes: bytes) -> ParseResult:
    """Fast extraction via pymupdf4llm — good text + basic tables."""
    import pymupdf4llm
    import pymupdf

    doc = pymupdf.Document(stream=pdf_bytes, filetype="pdf")
    page_count = len(doc)

    md_text = pymupdf4llm.to_markdown(doc)

    metadata = {}
    info = doc.metadata
    if info:
        for key in ("title", "author", "subject", "creator"):
            if info.get(key):
                metadata[key] = info[key]

    doc.close()

    return ParseResult(
        content=md_text,
        engine_used="pymupdf",
        page_count=page_count,
        metadata=metadata,
    )


def _parse_pdf_docling(pdf_bytes: bytes) -> ParseResult:
    """High-accuracy extraction via Docling — 97.9% table accuracy."""
    from docling.document_converter import DocumentConverter

    # Docling needs a file path, write to temp
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(pdf_bytes)
        tmp_path = tmp.name

    try:
        converter = DocumentConverter()
        result = converter.convert(tmp_path)
        md_text = result.document.export_to_markdown()

        return ParseResult(
            content=md_text,
            engine_used="docling",
            page_count=0,  # Docling doesn't expose page count easily
            metadata={},
        )
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def _parse_pdf_pypdf2(pdf_bytes: bytes) -> ParseResult:
    """Fallback: PyPDF2 text-only extraction."""
    from PyPDF2 import PdfReader

    reader = PdfReader(io.BytesIO(pdf_bytes))
    pages_text = [page.extract_text() or "" for page in reader.pages]
    content = "\n\n".join(pages_text)

    metadata = {}
    if reader.metadata:
        for key in ("title", "author", "subject", "creator"):
            val = getattr(reader.metadata, key, None)
            if val:
                metadata[key] = val

    return ParseResult(
        content=content,
        engine_used="pypdf2",
        page_count=len(reader.pages),
        metadata=metadata,
        warnings=["PyPDF2: tables and images not preserved"],
    )


def parse_pdf(
    pdf_bytes: bytes,
    engine: ParseEngine = ParseEngine.AUTO,
) -> ParseResult:
    """Parse a PDF file into Markdown text.

    Engine selection:
      - AUTO: pymupdf > docling > pypdf2 (best available)
      - Explicit: use the specified engine, fall back if unavailable
    """
    if engine == ParseEngine.AUTO:
        # Check config preference
        try:
            from core.config import load_config
            preferred = load_config().get("knowledge", {}).get("pdf_engine", "")
        except Exception:
            preferred = ""

        if preferred == "docling" and _has_docling():
            return _parse_pdf_docling(pdf_bytes)

        # Default: pymupdf > docling > pypdf2
        if _has_pymupdf():
            return _parse_pdf_pymupdf(pdf_bytes)
        if _has_docling():
            return _parse_pdf_docling(pdf_bytes)
        if _has_pypdf2():
            return _parse_pdf_pypdf2(pdf_bytes)
        raise ImportError(
            "No PDF parser available. Install one of: "
            "pymupdf4llm, docling, PyPDF2"
        )

    engine_map = {
        ParseEngine.PYMUPDF: (_has_pymupdf, _parse_pdf_pymupdf),
        ParseEngine.DOCLING: (_has_docling, _parse_pdf_docling),
        ParseEngine.PYPDF2: (_has_pypdf2, _parse_pdf_pypdf2),
    }

    check, parser = engine_map[engine]
    if check():
        return parser(pdf_bytes)

    # Fall back to AUTO
    logger.warning("Engine %s not available, falling back to AUTO", engine)
    return parse_pdf(pdf_bytes, engine=ParseEngine.AUTO)


# ── Excel Parsing ────────────────────────────────────────────

def parse_excel(
    file_bytes: bytes,
    filename: str = "file.xlsx",
    max_rows: int = 2000,
) -> ParseResult:
    """Parse Excel (.xlsx/.xls/.csv) into Markdown tables.

    For .xlsx: uses openpyxl first for structure + formula extraction,
    falls back to pandas for .xls/.csv or on error.
    """
    ext = Path(filename).suffix.lower()

    # CSV/TSV/XLS: pandas-only path
    if ext in (".csv", ".tsv", ".xls"):
        return _parse_excel_pandas(file_bytes, filename, ext, max_rows)

    # .xlsx: openpyxl with formula awareness
    try:
        return _parse_excel_openpyxl(file_bytes, max_rows)
    except Exception:
        logger.debug("openpyxl parse failed, falling back to pandas")
        return _parse_excel_pandas(file_bytes, filename, ext, max_rows)


def _parse_excel_openpyxl(file_bytes: bytes, max_rows: int) -> ParseResult:
    """Parse .xlsx with openpyxl — preserves sheet names and formulas."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
    sections: list[str] = []
    total_rows = 0
    warnings: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sections.append(f"## Sheet: {sheet_name}\n")

        formula_cells: list[str] = []
        rows_data: list[list[str]] = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=False)):
            if row_idx >= max_rows:
                warnings.append(f"Sheet '{sheet_name}' truncated to {max_rows} rows")
                break
            row_vals: list[str] = []
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cells.append(f"{cell.coordinate}: `{cell.value}`")
                    row_vals.append(cell.value)
                else:
                    row_vals.append(str(cell.value) if cell.value is not None else "")
            rows_data.append(row_vals)

        total_rows += len(rows_data)

        if rows_data:
            ncols = max(len(r) for r in rows_data)
            # Pad rows to same width
            for r in rows_data:
                while len(r) < ncols:
                    r.append("")
            header = "| " + " | ".join(rows_data[0]) + " |"
            sep = "| " + " | ".join(["---"] * ncols) + " |"
            body = "\n".join("| " + " | ".join(r) + " |" for r in rows_data[1:])
            sections.append(f"{header}\n{sep}\n{body}\n")

        if formula_cells:
            sections.append(f"\n**Formulas in {sheet_name}:**\n")
            for fc in formula_cells[:20]:
                sections.append(f"- {fc}")
            if len(formula_cells) > 20:
                sections.append(f"- ... and {len(formula_cells) - 20} more")
            sections.append("")

    return ParseResult(
        content="\n".join(sections),
        engine_used="openpyxl",
        page_count=len(wb.sheetnames),
        metadata={"sheets": wb.sheetnames, "total_rows": total_rows},
        warnings=warnings,
    )


def _parse_excel_pandas(
    file_bytes: bytes, filename: str, ext: str, max_rows: int
) -> ParseResult:
    """Fallback pandas parser for .xls/.csv/.tsv."""
    import pandas as pd

    warnings: list[str] = []

    if ext in (".csv", ".tsv"):
        df = pd.read_csv(io.BytesIO(file_bytes), nrows=max_rows)
        sheets = {"Sheet1": df}
    else:
        sheets = pd.read_excel(
            io.BytesIO(file_bytes), sheet_name=None, nrows=max_rows,
        )

    md_parts: list[str] = []
    total_rows = 0

    for sheet_name, df in sheets.items():
        total_rows += len(df)
        md_parts.append(f"## Sheet: {sheet_name}\n")
        md_parts.append(f"Rows: {len(df)}, Columns: {len(df.columns)}\n")

        if len(df) > max_rows:
            df = df.head(max_rows)
            warnings.append(f"Sheet '{sheet_name}' truncated to {max_rows} rows")

        md_parts.append(df.to_markdown(index=False))
        md_parts.append("")

    return ParseResult(
        content="\n".join(md_parts),
        engine_used="pandas",
        page_count=len(sheets),
        metadata={"total_rows": total_rows, "sheet_count": len(sheets)},
        warnings=warnings,
    )


# ── Unified entry point ─────────────────────────────────────

EXCEL_EXTENSIONS = {".xlsx", ".xls", ".csv", ".tsv"}
PDF_EXTENSIONS = {".pdf"}
TEXT_EXTENSIONS = {".txt", ".md", ".json", ".yaml", ".yml", ".xml", ".html"}


def parse_file(
    file_bytes: bytes,
    filename: str,
    engine: ParseEngine = ParseEngine.AUTO,
) -> ParseResult:
    """Parse any supported file type into Markdown text.

    Supported formats:
      - PDF: .pdf
      - Excel: .xlsx, .xls, .csv, .tsv
      - Text: .txt, .md, .json, .yaml, .xml, .html
    """
    ext = Path(filename).suffix.lower()

    if ext in PDF_EXTENSIONS:
        return parse_pdf(file_bytes, engine=engine)

    if ext in EXCEL_EXTENSIONS:
        return parse_excel(file_bytes, filename=filename)

    if ext in TEXT_EXTENSIONS:
        try:
            text = file_bytes.decode("utf-8")
        except UnicodeDecodeError:
            text = file_bytes.decode("utf-8", errors="replace")
        return ParseResult(
            content=text,
            engine_used="text",
            metadata={"encoding": "utf-8"},
        )

    # Unknown format — try as text
    try:
        text = file_bytes.decode("utf-8")
        return ParseResult(
            content=text,
            engine_used="text",
            warnings=[f"Unknown extension '{ext}', treated as text"],
        )
    except UnicodeDecodeError:
        raise ValueError(f"Unsupported file format: {ext}")
